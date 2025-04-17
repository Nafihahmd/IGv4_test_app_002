import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Define the XLSX file path.
TEST_REPORT = "test_results.xlsx"
MAC_LIST    = "mac_addr.xlsx"

# Define fill colors for pass and fail
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# Define a grey fill to mark a MAC address as used.
GREY_FILL = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")

def initialize_workbook():
    """Creates a new workbook with headers if it doesn't exist."""
    if not os.path.exists(TEST_REPORT):
        wb = Workbook()
        ws = wb.active
        ws.title = "TestResults"
        # Write headers: first column is a timestamp, then one per test.
        headers = ["Timestamp", "MAC Addr", "Ethernet", "USB", "RTC", "Xbee", "Battery", "Relay"]
        ws.append(headers)
        wb.save(TEST_REPORT)
        print("Created new workbook with headers.")
    else:
        print("Workbook already exists.")

def append_test_results(test_results, mac_addr):
    """
    Appends a new row to the Excel workbook.
    
    test_results: list of tuples. Each tuple is (test_name, result, detail)
    For this example we assume test_results is in the same order as the header columns.
    Only the result value is written and colored.
    """
    if not os.path.exists(TEST_REPORT):
        initialize_workbook()

    wb = load_workbook(TEST_REPORT)
    ws = wb.active
     # Create a timestamp for this row
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Define the order in which tests should appear; adjust as needed.
    test_order = ["Ethernet Test", "USB Test", "RTC Test", "Xbee Test", "Battery Test", "Relay Test"]
    
    # Construct the row: first element is timestamp then one cell for each test.
    row_data = [timestamp] + [mac_addr] +[test_results.get(test, "N/A") for test in test_order]
    ws.append(row_data)
    row_index = ws.max_row  # the new row's index

    # Apply colors based on the results (columns start at 2 because col 1 is the timestamp)
    for col_index, test in enumerate(test_order, start=2):
        cell = ws.cell(row=row_index, column=col_index)
        result = cell.value.strip().upper() if cell.value else ""
        if result == "FAIL":
            cell.fill = RED_FILL

    wb.save(TEST_REPORT)
    print(f"Test results appended to row {row_index}.")

def get_next_available_mac(mark_as_used=False):
    """
    Opens an Excel file containing MAC addresses and a corresponding status,
    and returns the first available MAC address (one whose status is not "gone!").
    It then updates that row's status to 'gone!' and fills the cell in grey
    to prevent duplication.
    
    Expected Excel format (first row is header):
        | MAC addr        | Status  |
      1 | 00:11:22:33:44:55 |         |
      2 | ...             | ...     |
    
    Returns:
        The available MAC address as a string, or None if no MAC addresses are available.
    """
    # Ensure that the file exists.
    if not os.path.exists(MAC_LIST):
        print(f"Error: {MAC_LIST} does not exist.")
        return None

    try:
        wb = load_workbook(MAC_LIST)
        ws = wb.active

        # Read the header row to find the columns for MAC address and Status.
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        mac_col_idx = None
        status_col_idx = None

        # Headers in the Excel file are assumed to be exactly "MAC addr" and "Status"
        for idx, header in enumerate(header_row):
            if header and str(header).strip().lower() == "mac addr":
                mac_col_idx = idx
            elif header and str(header).strip().lower() == "status":
                status_col_idx = idx

        if mac_col_idx is None or status_col_idx is None:
            print("Error: Could not find required columns ('MAC addr' and 'Status').")
            return None

        # Iterate over the rows starting from the second row.
        for row in ws.iter_rows(min_row=2):
            mac_cell = row[mac_col_idx]
            status_cell = row[status_col_idx]

            # Double-check that the status is either empty or not marked as 'gone!'.
            current_status = (status_cell.value or "").strip().lower()
            if current_status != "used":
                # Before returning, update the status.
                if mark_as_used:
                    status_cell.value = "used"
                    status_cell.fill = GREY_FILL  # grey out the cell
                else:
                    status_cell.value = "read"
                wb.save(MAC_LIST)
                print(f"MAC address {mac_cell.value} has been marked as {'used' if mark_as_used else 'read'}.")
                return mac_cell.value

        print("No available MAC address found.")
        return None

    except Exception as e:
        print("Error reading or updating the Excel file:", e)
        return None